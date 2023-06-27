import os
from docx import Document
from tqdm import tqdm
from datetime import datetime
from docx.shared import Pt
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT, WD_TAB_ALIGNMENT
from openpyxl import load_workbook
import re
import PySimpleGUI as sg

# Obter a data atual
data_atual = datetime.now().strftime("%d/%m/%Y")

def preencher_word():
    for i in tqdm(range(2,122), desc ="Carregando...", leave=False, ncols=70 ):
        def pasta(caminho):
            pasta = caminho
            #verificar se a pasta existe se não existir ele ira criar
            if not os.path.exists(pasta):
                os.makedirs(pasta)
                                            
        # Carregar o arquivo da planilha
        workbook = load_workbook(os.path.abspath('contratacao-UE_MS-14062023.xlsx'))
        
        # Selecionar a planilha desejada (por nome ou índice)
        sheet = workbook.worksheets[0]
        
        try:
            #conteudo da celula
            matricula = sheet.cell(row=i, column=1).value
            nome = sheet.cell(row=i, column=2).value
            cpf = sheet.cell(row=i, column=3).value
            lotacao = sheet.cell(row=i, column=6).value
            # Remover os números da lotação
            lotacao_sem_numeros = re.sub(r'\d+', '', lotacao)
            # Remover o primeiro traço (-)
            lotacao_sem_traco = lotacao_sem_numeros.split('-', 1)[-1].strip()
        except:
            break
        
        data_troca = sheet.cell(row=i, column=8).value
        rg = sheet.cell(row=i, column=9).value   
        orgao_emisor = sheet.cell(row=i, column=10).value                          
        data_emissao = sheet.cell(row=i, column=11).value
        
        # Converter as datas para o formato desejado
        if isinstance(data_troca, datetime):
            data_troca = data_troca.date().strftime('%d/%m/%Y')
        if isinstance(data_emissao, datetime):
            data_emissao = data_emissao.date().strftime('%d/%m/%Y')
            
        #criar pasta com nome da pessoa
        pasta(os.path.abspath(f'pessoas'))
        
        # Carregar o arquivo existente
        doc = Document(os.path.abspath('termo aditivo ACS p ACM.docx'))

        # Acessar o conteúdo do documento
        paragraphs = doc.paragraphs

        # Definir a formatação para Arial 12
        font_name = 'Arial'
        font_size = Pt(12)

        # Editar o conteúdo do parágrafo 5
        texto_1 = f'Eu, {nome}, CPF {cpf}, carteira de identificação nº {rg}, emitida em {data_emissao}, órgão emissor {orgao_emisor}, aprovado e classificado em Processo Seletivo Simplificado, para os trabalhos do CENSO DEMOGRÁFICO 2022, para exercer a função de AGENTE CENSITÁRIO SUPERVISOR, sob a matrícula {matricula}.'
        paragraphs[4].text = texto_1

        # Editar o conteúdo do parágrafo 6
        paragrafo_edicao = f'Declaro, para os devidos fins, e por livre e espontânea vontade, que aceito exercer a função de AGENTE CENSITÁRIO MUNICIPAL – ACM a partir do dia {data_troca}, tendo meu salário e minhas atribuições alterados conforme detalhamento dessa função descrito no referido Edital do Processo Seletivo Simplificado em questão.'
        paragraphs[6].text = paragrafo_edicao

        paragrafo_edicao = f'Data: {data_atual}'
        paragraphs[10].text = paragrafo_edicao
        
        # Aplicar formatação para Arial 12 em todos os parágrafos
        for paragraph in paragraphs:
            for run in paragraph.runs:
                run.font.name = font_name
                run.font.size = font_size
        
        # Definir o tamanho da fonte para o parágrafo 2 como 14
        for run in paragraphs[1].runs:
            run.font.name = font_name
            run.font.size = Pt(14)

        # Remover centralização e adicionar um recuo de tabulação ao parágrafo 1
        paragraph_1 = paragraphs[0]
        paragraph_1.alignment = WD_PARAGRAPH_ALIGNMENT.LEFT
        paragraph_1.paragraph_format.tab_stops.add_tab_stop(Pt(36), WD_TAB_ALIGNMENT.LEFT)
        
        # Definir o tamanho da fonte para o parágrafo 3 como 11
        for run in paragraphs[2].runs:
            run.font.name = font_name
            run.font.size = Pt(11)
        
        '''# Centralizar parágrafos 21 e 22
        paragraphs[20].alignment = WD_PARAGRAPH_ALIGNMENT.CENTER
        '''
        # Salvar as alterações
        doc.save(os.path.abspath(f'pessoas//{lotacao_sem_traco} - {nome} {i}.docx'))
        
        # Atualizar a barra de progresso na interface gráfica
        window['progress'].update(i + 1)
        window.refresh()
    
    sg.popup('Concluído!', title='Informação',keep_on_top=True)
        
selected_theme = 'Reddit'
sg.theme(selected_theme)        
layout_login = [
    [sg.ProgressBar(120, orientation='h', size=(20, 20), key='progress')],
    [sg.Stretch(),sg.Button('Ok',size=(5,1))]
    ]
        
window = sg.Window(' ', icon='favicon.ico',layout=layout_login, keep_on_top=True, finalize = True)

while True:
    event,values = window.read()
    if event in (None, 'Sair'):
        break
    
    if event == 'Ok':
        preencher_word()
        
window.close()